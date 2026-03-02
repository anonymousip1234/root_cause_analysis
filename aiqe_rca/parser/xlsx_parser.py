"""XLSX spreadsheet parser using openpyxl."""

import io
import uuid

from openpyxl import load_workbook

from aiqe_rca.models.evidence import EvidenceElement, SourceType


def parse_xlsx(filename: str, content: bytes) -> list[EvidenceElement]:
    """Parse an XLSX file into evidence elements.

    Reads each sheet row-by-row, using the first row as headers.
    Each data row becomes one evidence element.
    """
    elements: list[EvidenceElement] = []
    counter = 0

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # First row = headers
        headers = [str(h).strip() if h else f"Col{i+1}" for i, h in enumerate(rows[0])]

        for row_idx, row in enumerate(rows[1:], start=2):
            parts = []
            for col_idx, cell_val in enumerate(row):
                if cell_val is not None and str(cell_val).strip():
                    header = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx+1}"
                    parts.append(f"{header}: {cell_val}")
            row_text = "; ".join(parts)
            if len(row_text) < 10:
                continue
            counter += 1
            elements.append(
                EvidenceElement(
                    id=f"E-{uuid.uuid5(uuid.NAMESPACE_DNS, f'{filename}-{sheet_name}-r{row_idx}')}",
                    source=filename,
                    source_type=SourceType.XLSX,
                    text_content=row_text,
                    page_ref=f"sheet: {sheet_name}, row {row_idx}",
                )
            )

    wb.close()
    return elements
